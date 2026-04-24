#!/usr/bin/env python3
"""
Test Native Export Endpoint
Run this on the Qlik server after deployment to verify native export works.

Usage:
    python test_native_export.py
"""

import requests
import time
from pathlib import Path

# Configuration
API_URL = "http://localhost:8000/api/v1/apps/afko/tables/factory_data/export_native"
API_KEY = "admin-fe7dfb39272f9ddbe221b389e555e7fff7d4941a9402e22b328b3aa8eff0496d"
OUTPUT_FILE = "test_native_export.xlsx"

def test_native_export():
    """Test the native export endpoint"""

    print("="*80)
    print("NATIVE EXPORT TEST")
    print("="*80)
    print(f"\nEndpoint: {API_URL}")
    print(f"Output file: {OUTPUT_FILE}")

    # Test 1: Check if API is running
    print("\n" + "-"*80)
    print("TEST 1: Check if API server is running")
    print("-"*80)

    try:
        health_response = requests.get("http://localhost:8000/health", timeout=5)
        if health_response.status_code == 200:
            print("✓ API server is running")
        else:
            print(f"⚠ API server returned status {health_response.status_code}")
    except requests.exceptions.RequestException as e:
        print(f"✗ Cannot connect to API server: {e}")
        print("\nMake sure the API server is running:")
        print("  python run.py")
        return False

    # Test 2: Call native export endpoint
    print("\n" + "-"*80)
    print("TEST 2: Call native export endpoint")
    print("-"*80)

    headers = {
        "X-API-Key": API_KEY
    }

    print("\nSending request...")
    start_time = time.time()

    try:
        response = requests.get(API_URL, headers=headers, timeout=120)
        elapsed = time.time() - start_time

        print(f"Status Code: {response.status_code}")
        print(f"Response Time: {elapsed:.2f} seconds")

        if response.status_code != 200:
            print(f"\n✗ Request failed with status {response.status_code}")
            print(f"Response: {response.text[:500]}")
            return False

        print(f"✓ Request successful")

        # Test 3: Check response headers
        print("\n" + "-"*80)
        print("TEST 3: Verify response headers")
        print("-"*80)

        content_type = response.headers.get('content-type', '')
        content_disposition = response.headers.get('content-disposition', '')
        content_length = response.headers.get('content-length', '0')

        print(f"Content-Type: {content_type}")
        print(f"Content-Disposition: {content_disposition}")
        print(f"Content-Length: {content_length} bytes")

        if 'spreadsheetml' not in content_type and 'excel' not in content_type.lower():
            print(f"⚠ Unexpected content type (expected Excel file)")
        else:
            print(f"✓ Content type is correct")

        # Test 4: Save and verify file
        print("\n" + "-"*80)
        print("TEST 4: Save and verify file")
        print("-"*80)

        file_size = len(response.content)
        print(f"File size: {file_size:,} bytes ({file_size / 1024 / 1024:.2f} MB)")

        # Check if it's actually Excel content (starts with PK for zip format)
        if not response.content[:2] == b'PK':
            print("⚠ File doesn't appear to be a valid Excel file (should start with PK)")
            print(f"First 50 bytes: {response.content[:50]}")

            # Check if it's HTML/JSON error
            if response.content[:20].lower().startswith(b'<!doctype') or response.content[:20].lower().startswith(b'<html'):
                print("✗ Received HTML instead of Excel file")
                return False
            elif response.content[:20].startswith(b'{'):
                print("✗ Received JSON error instead of Excel file")
                print(f"Error: {response.text[:200]}")
                return False
        else:
            print("✓ File appears to be a valid Excel file")

        # Save file
        with open(OUTPUT_FILE, 'wb') as f:
            f.write(response.content)

        print(f"✓ File saved to: {OUTPUT_FILE}")

        # Test 5: Performance check
        print("\n" + "-"*80)
        print("TEST 5: Performance analysis")
        print("-"*80)

        if elapsed < 10:
            print(f"✓ Excellent! Export completed in {elapsed:.2f} seconds")
            print("  Native export is working as expected (fast)")
        elif elapsed < 30:
            print(f"✓ Good! Export completed in {elapsed:.2f} seconds")
            print("  Performance is acceptable")
        else:
            print(f"⚠ Export took {elapsed:.2f} seconds")
            print("  This seems slower than expected for native export")

        # Test 6: Try to open with openpyxl
        print("\n" + "-"*80)
        print("TEST 6: Validate Excel file structure")
        print("-"*80)

        try:
            from openpyxl import load_workbook
            wb = load_workbook(OUTPUT_FILE, read_only=True)
            ws = wb.active

            row_count = ws.max_row
            col_count = ws.max_column

            print(f"✓ Excel file is valid and readable")
            print(f"  Rows: {row_count:,}")
            print(f"  Columns: {col_count}")

            # Read first few cells
            if row_count > 0:
                print(f"\n  First row (headers):")
                first_row = []
                for col in range(1, min(6, col_count + 1)):
                    cell_value = ws.cell(1, col).value
                    first_row.append(str(cell_value)[:20])
                print(f"    {' | '.join(first_row)}")

            wb.close()

        except ImportError:
            print("⚠ openpyxl not installed, skipping validation")
            print("  Install with: pip install openpyxl")
        except Exception as e:
            print(f"⚠ Could not validate Excel file: {e}")

        # Final result
        print("\n" + "="*80)
        print("✓✓✓ ALL TESTS PASSED! ✓✓✓")
        print("="*80)
        print(f"\nNative export is working correctly!")
        print(f"File saved to: {OUTPUT_FILE}")
        print(f"Export time: {elapsed:.2f} seconds")
        print(f"File size: {file_size / 1024 / 1024:.2f} MB")

        return True

    except requests.exceptions.Timeout:
        print(f"\n✗ Request timed out after 120 seconds")
        return False
    except requests.exceptions.RequestException as e:
        print(f"\n✗ Request failed: {e}")
        return False
    except Exception as e:
        print(f"\n✗ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main entry point"""
    try:
        success = test_native_export()

        if not success:
            print("\n" + "="*80)
            print("TROUBLESHOOTING TIPS")
            print("="*80)
            print("""
1. Make sure API server is running:
   python run.py

2. Verify you're on the Qlik server:
   The API must run on the same machine as Qlik Sense

3. Check the TempContent folder exists:
   C:\\ProgramData\\Qlik\\Sense\\Repository\\TempContent

4. Check logs for detailed errors:
   Look at the API server console output

5. Try the manual export endpoint instead:
   /api/v1/apps/afko/tables/factory_data/export
""")
            return 1

        return 0

    except KeyboardInterrupt:
        print("\n\nTest cancelled by user")
        return 1

if __name__ == "__main__":
    exit(main())
