from fastapi import APIRouter, Depends, Path, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from src.api.services.app_service import AppService
from src.api.core.dependencies import get_app_service, verify_api_key
from src.api.core.config import settings
from src.api.clients.qlik_engine import QlikEngineClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import logging

logger = logging.getLogger(__name__)
router = APIRouter()

# IMPORTANT: More specific routes must come BEFORE generic routes
# These specific routes must be defined before the generic {table_name} route

@router.get("/apps/{app_name}/tables/factory_data/data")
async def get_factory_data(
    app_name: str = Path(..., description="Application name"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=10000, description="Rows per page"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Get factory data with optional filtering and variable controls.

    This endpoint retrieves data from the factory_data pivot table (object ID: Dkjpv)
    and supports filtering by Завод (factory) and Склад (warehouse) fields,
    plus setting variables for measure type and currency.

    **Examples:**

    Get all data (no filtering):
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100
    ```

    Filter by single factory:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&factory=1203
    ```

    Filter by multiple factories:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&factory=1203,1204
    ```

    Filter by warehouse:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&warehouse=A100
    ```

    Filter by OM type:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&typeOM=Type1
    ```

    Filter by factory, warehouse, and OM type with variables:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&factory=1203&warehouse=A100&typeOM=Type1&MeasureType=1&Currency=2
    ```

    **Response format:**
    ```json
    {
      "object_id": "Dkjpv",
      "app_id": "...",
      "app_name": "afko",
      "data": [
        {
          "Field1": "Value1",
          "Field2": "Value2",
          ...
        }
      ],
      "pagination": {
        "page": 1,
        "page_size": 100,
        "total_rows": 150,
        "total_pages": 2,
        "has_next": true,
        "has_previous": false
      }
    }
    ```
    """
    table_name = "factory_data"

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Get bookmark ID for this table (applies 3-month filter first)
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # Build selections dictionary for Qlik filtering
    # NOTE: yearMonth is NOT included in selections - it will be filtered client-side
    selections = {}
    if factory:
        # Split comma-separated values
        factory_values = [f.strip() for f in factory.split(',')]
        selections['Завод'] = factory_values  # Завод is the factory field in Qlik

    if warehouse:
        # Split comma-separated values
        warehouse_values = [w.strip() for w in warehouse.split(',')]
        selections['Склад'] = warehouse_values  # Склад is the warehouse field in Qlik

    if typeOM:
        # Split comma-separated values
        typeOM_values = [t.strip() for t in typeOM.split(',')]
        selections['Тип ОМ'] = typeOM_values  # Тип ОМ is the OM type field in Qlik

    # Build filters dictionary for client-side filtering
    # yearMonth will be filtered by parsing the Дата (Date) field after fetching from Qlik
    filters = {}
    if yearMonth:
        # Parse yearMonth values for client-side filtering
        # Input format: "2024-01" or "2024.01"
        year_month_list = []
        for ym in yearMonth.split(','):
            ym = ym.strip()
            # Normalize to YYYY-MM format
            ym_normalized = ym.replace('.', '-')
            year_month_list.append(ym_normalized)
        filters['yearMonth'] = year_month_list

    # Build variables dictionary
    variables = {}
    if MeasureType:
        variables['vChooseType'] = MeasureType
    if Currency:
        variables['vChooseCur'] = Currency

    # Fetch data with bookmark applied first (filters to 3 months), then apply selections
    # yearMonth is filtered client-side by parsing the Дата field
    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=page,
        page_size=page_size,
        filters=filters,  # Client-side filtering (yearMonth by Дата field)
        selections=selections,  # Apply Qlik selections for filtering
        variables=variables,  # Apply Qlik variables
        bookmark_id=bookmark_id  # Apply bookmark FIRST to filter to 3 months
    )

    return data


@router.get("/apps/{app_name}/tables/factory_data/export")
async def export_factory_data_to_excel(
    app_name: str = Path(..., description="Application name"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Export factory data to Excel file.

    This endpoint exports all factory data (with applied filters) to an Excel file.
    The Excel file includes formatted headers and all data rows.

    **Examples:**

    Export all data for factory 1203:
    ```
    GET /api/v1/apps/afko/tables/factory_data/export?factory=1203
    ```

    Export with multiple filters:
    ```
    GET /api/v1/apps/afko/tables/factory_data/export?factory=1203&yearMonth=2026-03&typeOM=Отгрузка в РЦ
    ```
    """

    # Get bookmark ID for this table
    table_name = "factory_data"
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # Build selections dictionary (same as JSON endpoint)
    selections = {}
    if factory:
        factory_values = [f.strip() for f in factory.split(',')]
        selections['Завод'] = factory_values

    if warehouse:
        warehouse_values = [w.strip() for w in warehouse.split(',')]
        selections['Склад'] = warehouse_values

    if typeOM:
        typeOM_values = [t.strip() for t in typeOM.split(',')]
        selections['Тип ОМ'] = typeOM_values

    # Build filters dictionary for client-side filtering
    filters = {}
    if yearMonth:
        yearMonth_values = [ym.strip().replace('.', '-') for ym in yearMonth.split(',')]
        filters['yearMonth'] = yearMonth_values

    # Force full dimension expansion for Excel export
    # Set a dummy filter that will force session hypercube creation
    # This ensures we always get all 7 dimensions, not just the 5 visible in pivot
    filters['_force_session_hypercube'] = True

    # Build variables dictionary
    variables = {}
    if MeasureType:
        variables['vChooseType'] = MeasureType
    if Currency:
        variables['vChooseCur'] = Currency

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # For export, we want ALL data without pagination limits
    # Set page_size to a very large number to ensure we get everything
    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=1,
        page_size=999999999,  # Effectively unlimited - get ALL rows
        filters=filters,
        selections=selections,
        variables=variables,
        bookmark_id=bookmark_id
    )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Factory Data"

    # Get data rows
    rows = data.get('data', [])

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for the given filters")

    # Get column headers from first row
    headers = list(rows[0].keys())

    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"factory_data_{timestamp}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/apps/{app_name}/tables/factory_data/export_native")
async def export_factory_data_native(
    app_name: str = Path(..., description="Application name"),
    file_type: str = Query("excel", description="Export format: excel, csv, tsv, or parquet"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    api_key: str = Depends(verify_api_key)
):
    """
    Export factory data using Qlik's native ExportData method (MUCH FASTER).

    This endpoint uses Qlik Sense's built-in ExportData API which can export
    up to 1 million rows directly to Excel, CSV, TSV, or Parquet format.

    **Supported formats:**
    - excel (default): Excel .xlsx format
    - csv: Comma-separated values
    - tsv: Tab-separated values
    - parquet: Apache Parquet format (requires Qlik Sense November 2024+ and app-level PARQUET support)

    **Note:** If PARQUET format is not supported by your Qlik version/app, it will automatically
    fall back to Excel format.

    **This is significantly faster than the regular export endpoint for large datasets.**
    """
    # Map file_type parameter to Qlik format codes
    format_mapping = {
        "excel": ("OOXML", ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "csv": ("CSV_C", ".csv", "text/csv"),
        "tsv": ("CSV_T", ".tsv", "text/tab-separated-values"),
        "parquet": ("PARQUET", ".parquet", "application/octet-stream")
    }

    if file_type not in format_mapping:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file_type '{file_type}'. Supported: excel, csv, tsv, parquet"
        )

    qlik_format, file_extension, media_type = format_mapping[file_type]
    table_name = "factory_data_table"  # Use the flat table object, not pivot

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access (use factory_data for access control)
    if not settings.can_access_table(api_key, app_name, "factory_data"):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to factory_data table in app '{app_name}'"
        )

    # Get object ID for the table object
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Get app ID
    app_id = settings.get_app_id(app_name)
    if not app_id:
        raise HTTPException(
            status_code=404,
            detail=f"App '{app_name}' not found in configuration"
        )

    # Get bookmark ID
    bookmark_id = settings.get_bookmark_id(app_name, "factory_data")

    # Create Qlik Engine client
    client = QlikEngineClient(settings)

    try:
        # Connect and open app
        client.connect()
        result = client.open_doc(app_id, no_data=False)
        app_handle = result['qReturn']['qHandle']

        # Apply bookmark if specified
        if bookmark_id:
            client.send_request('ApplyBookmark', [bookmark_id], handle=app_handle)

        # Apply variables if specified (must be done before getting object)
        if MeasureType:
            client.set_variable_value(app_handle, 'vChooseType', MeasureType)

        if Currency:
            client.set_variable_value(app_handle, 'vChooseCur', Currency)

        # Apply field selections for filtering
        if factory:
            factory_values = [f.strip() for f in factory.split(',')]
            client.select_in_field(app_handle, 'Завод', factory_values, toggle=False)

        if warehouse:
            warehouse_values = [w.strip() for w in warehouse.split(',')]
            client.select_in_field(app_handle, 'Склад', warehouse_values, toggle=False)

        if typeOM:
            typeOM_values = [t.strip() for t in typeOM.split(',')]
            client.select_in_field(app_handle, 'Тип ОМ', typeOM_values, toggle=False)

        if yearMonth:
            yearMonth_values = [ym.strip().replace('.', '-') for ym in yearMonth.split(',')]
            client.select_in_field(app_handle, 'YearMonth', yearMonth_values, toggle=False)

        # Get object handle (after selections are applied)
        obj_result = client.send_request('GetObject', [object_id], handle=app_handle)
        obj_handle = obj_result['qReturn']['qHandle']

        # Use native ExportData method with fallback for unsupported formats
        export_result = None
        actual_format = qlik_format
        actual_extension = file_extension
        actual_media_type = media_type

        try:
            export_result = client.export_data(
                object_handle=obj_handle,
                file_type=qlik_format,
                path="/qHyperCubeDef",
                export_state="P"  # P = Possible values (respects current selections/filters)
            )
        except Exception as export_error:
            # Check if error is "Unsupported file format" (code 3004) for PARQUET
            if qlik_format == "PARQUET" and "3004" in str(export_error):
                logger.warning(
                    f"PARQUET format not supported (error 3004). "
                    f"Falling back to Excel format. "
                    f"To enable PARQUET: Add 'SET EnableParquetSupport=1;' to app load script."
                )
                # Fallback to Excel
                actual_format = "OOXML"
                actual_extension = ".xlsx"
                actual_media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                export_result = client.export_data(
                    object_handle=obj_handle,
                    file_type=actual_format,
                    path="/qHyperCubeDef",
                    export_state="P"  # P = Possible values (respects current selections/filters)
                )
            else:
                # Re-raise if it's a different error
                raise

        # Get the temporary URL
        temp_url = export_result.get('qUrl')
        if not temp_url:
            raise HTTPException(
                status_code=500,
                detail="Qlik did not return a download URL"
            )

        # Parse the file path from qUrl
        # Format: /tempcontent/GUID1/GUID2.xlsx?serverNodeId=...
        from pathlib import Path
        parts = temp_url.strip('/').split('/')
        guid_folder = parts[1]
        filename_with_params = parts[2]
        filename = filename_with_params.split('?')[0]  # Remove query params

        # Direct filesystem access (when API runs on same server as Qlik)
        # Standard Qlik TempContent location on Windows
        base_path = Path("C:/ProgramData/Qlik/Sense/Repository/TempContent")
        file_path = base_path / guid_folder / filename

        # Check if file exists
        if not file_path.exists():
            raise HTTPException(
                status_code=500,
                detail=f"Exported file not found at {file_path}. Ensure API runs on Qlik server."
            )

        # Read file directly from filesystem
        with open(file_path, 'rb') as f:
            file_content = f.read()

        # Generate filename for download
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"factory_data_{timestamp}{actual_extension}"

        # Return the file as streaming response
        return StreamingResponse(
            BytesIO(file_content),
            media_type=actual_media_type,
            headers={"Content-Disposition": f"attachment; filename={download_filename}"}
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Export failed: {str(e)}"
        )
    finally:
        # Close connection
        try:
            client.close()
        except:
            pass


@router.get("/apps/{app_name}/tables/factory_material_remainder/data")
async def get_factory_material_remainder(
    app_name: str = Path(..., description="Application name"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=10000, description="Rows per page"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Get factory material remainder data (materials remaining at end of day) with optional filtering.

    This endpoint retrieves data from the factory_material_remainder table (object ID: 26bfee13-02c5-41d6-afe7-439ec9545088)
    and supports filtering by Завод (factory), Склад (warehouse), and Тип ОМ fields,
    plus setting variables for measure type and currency.

    **Examples:**

    Get all data (no filtering):
    ```
    GET /api/v1/apps/afko/tables/factory_material_remainder/data?page=1&page_size=100
    ```

    Filter by factory:
    ```
    GET /api/v1/apps/afko/tables/factory_material_remainder/data?page=1&page_size=100&factory=1203
    ```

    Filter by multiple factories:
    ```
    GET /api/v1/apps/afko/tables/factory_material_remainder/data?page=1&page_size=100&factory=1203,1204
    ```
    """
    table_name = "factory_material_remainder"

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Get bookmark ID for this table
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # Build selections dictionary for Qlik filtering
    selections = {}
    if factory:
        factory_values = [f.strip() for f in factory.split(',')]
        selections['Завод'] = factory_values

    if warehouse:
        warehouse_values = [w.strip() for w in warehouse.split(',')]
        selections['Склад'] = warehouse_values

    if typeOM:
        typeOM_values = [t.strip() for t in typeOM.split(',')]
        selections['Тип ОМ'] = typeOM_values

    # Build filters dictionary for client-side filtering
    filters = {}
    if yearMonth:
        year_month_list = []
        for ym in yearMonth.split(','):
            ym = ym.strip()
            ym_normalized = ym.replace('.', '-')
            year_month_list.append(ym_normalized)
        filters['yearMonth'] = year_month_list

    # Build variables dictionary
    variables = {}
    if MeasureType:
        variables['vChooseType'] = MeasureType
    if Currency:
        variables['vChooseCur'] = Currency

    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=page,
        page_size=page_size,
        filters=filters,
        selections=selections,
        variables=variables,
        bookmark_id=bookmark_id
    )

    return data


@router.get("/apps/{app_name}/tables/factory_material_remainder/export")
async def export_factory_material_remainder_to_excel(
    app_name: str = Path(..., description="Application name"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Export factory material remainder data to Excel file.

    **Examples:**

    Export all data for factory 1203:
    ```
    GET /api/v1/apps/afko/tables/factory_material_remainder/export?factory=1203
    ```

    Export with multiple filters:
    ```
    GET /api/v1/apps/afko/tables/factory_material_remainder/export?factory=1203&yearMonth=2026-03
    ```
    """
    table_name = "factory_material_remainder"
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # Build selections dictionary
    selections = {}
    if factory:
        factory_values = [f.strip() for f in factory.split(',')]
        selections['Завод'] = factory_values

    if warehouse:
        warehouse_values = [w.strip() for w in warehouse.split(',')]
        selections['Склад'] = warehouse_values

    if typeOM:
        typeOM_values = [t.strip() for t in typeOM.split(',')]
        selections['Тип ОМ'] = typeOM_values

    # Build filters dictionary for client-side filtering
    filters = {}
    if yearMonth:
        yearMonth_values = [ym.strip().replace('.', '-') for ym in yearMonth.split(',')]
        filters['yearMonth'] = yearMonth_values

    # Force full dimension expansion for Excel export (pivot object needs this)
    filters['_force_session_hypercube'] = True

    # Build variables dictionary
    variables = {}
    if MeasureType:
        variables['vChooseType'] = MeasureType
    if Currency:
        variables['vChooseCur'] = Currency

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=1,
        page_size=999999999,
        filters=filters,
        selections=selections,
        variables=variables,
        bookmark_id=bookmark_id
    )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Factory Material Remainder"

    rows = data.get('data', [])

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for the given filters")

    headers = list(rows[0].keys())

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"factory_material_remainder_{timestamp}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/apps/{app_name}/tables/factory_material_remainder/export_native")
async def export_factory_material_remainder_native(
    app_name: str = Path(..., description="Application name"),
    file_type: str = Query("excel", description="Export format: excel, csv, tsv, or parquet"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    api_key: str = Depends(verify_api_key)
):
    """
    Export factory material remainder data using Qlik's native ExportData method (MUCH FASTER).

    This endpoint uses Qlik Sense's built-in ExportData API which can export
    up to 1 million rows directly to Excel, CSV, TSV, or Parquet format.

    **Supported formats:**
    - excel (default): Excel .xlsx format
    - csv: Comma-separated values
    - tsv: Tab-separated values
    - parquet: Apache Parquet format (requires Qlik Sense November 2024+ and app-level PARQUET support)

    **Note:** If PARQUET format is not supported by your Qlik version/app, it will automatically
    fall back to Excel format.

    **This is significantly faster than the regular export endpoint for large datasets.**
    """
    # Map file_type parameter to Qlik format codes
    format_mapping = {
        "excel": ("OOXML", ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "csv": ("CSV_C", ".csv", "text/csv"),
        "tsv": ("CSV_T", ".tsv", "text/tab-separated-values"),
        "parquet": ("PARQUET", ".parquet", "application/octet-stream")
    }

    if file_type not in format_mapping:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file_type '{file_type}'. Supported: excel, csv, tsv, parquet"
        )

    qlik_format, file_extension, media_type = format_mapping[file_type]
    table_name = "factory_material_remainder_table"  # Use the flat table object, not pivot

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access (use factory_material_remainder for access control)
    if not settings.can_access_table(api_key, app_name, "factory_material_remainder"):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to factory_material_remainder table in app '{app_name}'"
        )

    # Get object ID for the table object
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Get app ID
    app_id = settings.get_app_id(app_name)
    if not app_id:
        raise HTTPException(
            status_code=404,
            detail=f"App '{app_name}' not found in configuration"
        )

    # Get bookmark ID
    bookmark_id = settings.get_bookmark_id(app_name, "factory_material_remainder")

    # Create Qlik Engine client
    client = QlikEngineClient(settings)

    try:
        # Connect and open app
        client.connect()
        result = client.open_doc(app_id, no_data=False)
        app_handle = result['qReturn']['qHandle']

        # Apply bookmark if specified
        if bookmark_id:
            client.send_request('ApplyBookmark', [bookmark_id], handle=app_handle)

        # Apply variables if specified (must be done before getting object)
        if MeasureType:
            client.set_variable_value(app_handle, 'vChooseType', MeasureType)

        if Currency:
            client.set_variable_value(app_handle, 'vChooseCur', Currency)

        # Apply field selections for filtering
        if factory:
            factory_values = [f.strip() for f in factory.split(',')]
            client.select_in_field(app_handle, 'Завод', factory_values, toggle=False)

        if warehouse:
            warehouse_values = [w.strip() for w in warehouse.split(',')]
            client.select_in_field(app_handle, 'Склад', warehouse_values, toggle=False)

        if typeOM:
            typeOM_values = [t.strip() for t in typeOM.split(',')]
            client.select_in_field(app_handle, 'Тип ОМ', typeOM_values, toggle=False)

        if yearMonth:
            yearMonth_values = [ym.strip().replace('.', '-') for ym in yearMonth.split(',')]
            client.select_in_field(app_handle, 'YearMonth', yearMonth_values, toggle=False)

        # Get object handle (after selections are applied)
        obj_result = client.send_request('GetObject', [object_id], handle=app_handle)
        obj_handle = obj_result['qReturn']['qHandle']

        # Reorder dimensions so 'Дата' is the last dimension.
        # Master measures use GetObjectDimension(Dimensionality()-1) = 'Дата',
        # which requires 'Дата' to be the last dimension to return values.
        try:
            props_result = client.send_request('GetProperties', [], handle=obj_handle)
            props = props_result.get('qProp', {})
            cube_def = props.get('qHyperCubeDef', {})
            dimensions = list(cube_def.get('qDimensions', []))

            date_dim_idx = None
            for i, dim in enumerate(dimensions):
                field_defs = dim.get('qDef', {}).get('qFieldDefs', [])
                if any('Дата' in fd for fd in field_defs):
                    date_dim_idx = i
                    break

            if date_dim_idx is not None and date_dim_idx != len(dimensions) - 1:
                date_dim = dimensions.pop(date_dim_idx)
                dimensions.append(date_dim)
                cube_def['qDimensions'] = dimensions
                props['qHyperCubeDef'] = cube_def
                session_result = client.send_request('CreateSessionObject', [props], handle=app_handle)
                if 'qReturn' in session_result and 'qHandle' in session_result['qReturn']:
                    obj_handle = session_result['qReturn']['qHandle']
                    logger.info("Using session object with 'Дата' as last dimension for export")
        except Exception as reorder_error:
            logger.warning(f"Could not reorder dimensions for export: {reorder_error}, using original object")

        # Use native ExportData method with fallback for unsupported formats
        export_result = None
        actual_format = qlik_format
        actual_extension = file_extension
        actual_media_type = media_type

        try:
            export_result = client.export_data(
                object_handle=obj_handle,
                file_type=qlik_format,
                path="/qHyperCubeDef",
                export_state="P"  # P = Possible values (respects current selections/filters)
            )
        except Exception as export_error:
            # Check if error is "Unsupported file format" (code 3004) for PARQUET
            if qlik_format == "PARQUET" and "3004" in str(export_error):
                logger.warning(
                    f"PARQUET format not supported (error 3004). "
                    f"Falling back to Excel format. "
                    f"To enable PARQUET: Add 'SET EnableParquetSupport=1;' to app load script."
                )
                # Fallback to Excel
                actual_format = "OOXML"
                actual_extension = ".xlsx"
                actual_media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                export_result = client.export_data(
                    object_handle=obj_handle,
                    file_type=actual_format,
                    path="/qHyperCubeDef",
                    export_state="P"  # P = Possible values (respects current selections/filters)
                )
            else:
                # Re-raise if it's a different error
                raise

        # Get the temporary URL
        temp_url = export_result.get('qUrl')
        if not temp_url:
            raise HTTPException(
                status_code=500,
                detail="Qlik did not return a download URL"
            )

        # Parse the file path from qUrl
        # Format: /tempcontent/GUID1/GUID2.xlsx?serverNodeId=...
        from pathlib import Path
        parts = temp_url.strip('/').split('/')
        guid_folder = parts[1]
        filename_with_params = parts[2]
        filename = filename_with_params.split('?')[0]  # Remove query params

        # Direct filesystem access (when API runs on same server as Qlik)
        # Standard Qlik TempContent location on Windows
        base_path = Path("C:/ProgramData/Qlik/Sense/Repository/TempContent")
        file_path = base_path / guid_folder / filename

        # Check if file exists
        if not file_path.exists():
            raise HTTPException(
                status_code=500,
                detail=f"Exported file not found at {file_path}. Ensure API runs on Qlik server."
            )

        # Read file directly from filesystem
        with open(file_path, 'rb') as f:
            file_content = f.read()

        # Generate filename for download
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"factory_material_remainder_{timestamp}{actual_extension}"

        # Return the file as streaming response
        return StreamingResponse(
            BytesIO(file_content),
            media_type=actual_media_type,
            headers={"Content-Disposition": f"attachment; filename={download_filename}"}
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Export failed: {str(e)}"
        )
    finally:
        # Close connection
        try:
            client.close()
        except:
            pass


@router.get("/apps/{app_name}/tables/application_status/data")
async def get_application_status_data(
    app_name: str = Path(..., description="Application name"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=10000, description="Rows per page"),
    yearmonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2026.01 or 2026-01), supports multiple values separated by comma"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Get application status data with optional YearMonth filtering.

    This endpoint retrieves data from the application_status table and supports
    filtering by YearMonth field using Qlik selections. The YearMonth field
    exists in the app but may not be in the table itself.

    **Examples:**

    Get all data (no filtering):
    ```
    GET /api/v1/apps/Stock/tables/application_status/data?page=1&page_size=100
    ```

    Filter by single month:
    ```
    GET /api/v1/apps/Stock/tables/application_status/data?page=1&page_size=100&yearmonth=2024-01
    ```

    Filter by multiple months:
    ```
    GET /api/v1/apps/Stock/tables/application_status/data?page=1&page_size=100&yearmonth=2024-01,2024-02,2024-03
    ```

    **Response format:**
    ```json
    {
      "object_id": "UWDJj",
      "app_id": "...",
      "app_name": "Stock",
      "data": [
        {
          "Field1": "Value1",
          "Field2": "Value2",
          ...
        }
      ],
      "pagination": {
        "page": 1,
        "page_size": 100,
        "total_rows": 150,
        "total_pages": 2,
        "has_next": true,
        "has_previous": false
      }
    }
    ```
    """
    table_name = "application_status"

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Get bookmark ID for this table
    # The bookmark pre-filters the data to 3 months for fast retrieval
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # application_status uses the regular table object (UWDJj)
    # Apply bookmark FIRST to filter the data before creating the hypercube
    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=page,
        page_size=page_size,
        filters={},  # No client-side filtering
        selections={},  # No Qlik selections - rely on bookmark
        variables={},  # No variables for this endpoint
        bookmark_id=bookmark_id  # Applied FIRST to filter the data
    )

    return data


@router.get("/apps/{app_name}/tables/{table_name}/data")
async def get_table_data_with_measures(
    app_name: str = Path(..., description="Application name"),
    table_name: str = Path(..., description="Table name (e.g., stock_qty)"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=10000, description="Rows per page"),
    factory: Optional[str] = Query(None, description="Filter by factory (PRCTR field)"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (LGORT field)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Get actual data from a table with dimensions and measures.

    This endpoint retrieves actual data rows where each row contains
    all dimension values and calculated measure values.

    **Example:**
    ```
    GET /api/v1/apps/Stock/tables/stock_qty/data?page=1&page_size=10
    ```

    **With filters:**
    ```
    GET /api/v1/apps/Stock/tables/stock_qty/data?page=1&page_size=10&factory=1203&warehouse=P210
    ```

    **Response format:**
    ```json
    {
      "data": [
        {
          "MATNR": "000000001000000000",
          "Название материалов": "Замес д. ПВХ профилей PE",
          "База Код": "1203",
          "Название завода": "Производство ПВХ профилей",
          "Склад": "P210",
          "СвобИспользЗапас": 234.22,
          "Базовая ЕИ": "KG"
        }
      ],
      "pagination": {...}
    }
    ```
    """
    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Build selections dictionary for Qlik-side filtering (server-side, efficient pagination)
    # Map query parameters to actual Qlik field names
    selections = {}
    if factory:
        selections['PRCTR'] = [factory]  # PRCTR is the factory field in Qlik
    if warehouse:
        selections['LGORT'] = [warehouse]  # LGORT is the warehouse field in Qlik

    data = await app_service.get_object_data(app_name, object_id, page, page_size, filters={}, selections=selections, variables={})
    return data
